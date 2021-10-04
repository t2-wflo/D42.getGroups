from openpyxl import Workbook
import requests
import urllib3
import string


## Disable ssl warning
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

##Prep excel.
wb = Workbook()
ws = wb.active


## Device42 Credentials and API url.
class D42Creds():
	username = "username"
	password = "password"
	url = "https://device42.take2games.com/api/1.0/admingroups/"

d42Post = (requests.get(D42Creds.url, verify=False, auth=(D42Creds.username, D42Creds.password))).json()

columnNum = 1
for group in d42Post["admingroups"]:
	rowNum = 1
	ws.cell(row=1, column=columnNum, value=group["name"])
	for users in group["members"]:
		ws.cell(row=rowNum+1, column=columnNum, value=users)
		rowNum += 1
	columnNum +=1

# Save the file
wb.save("Groups.xlsx")
